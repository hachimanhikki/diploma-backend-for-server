import os
import openpyxl
from django.conf import settings
from api.model.models import Group, Subject
from config.settings import DepartmentEnum
import api.service.functions as functions
import collections


wb = course_sheet = course_educational_programs_count = group_sheet = schedule_wb = None


def set_workload() -> None:
    global wb, course_sheet, course_educational_programs_count, group_sheet
    wb = openpyxl.load_workbook(os.path.join(
        settings.MEDIA_ROOT, 'all_data.xlsx'), data_only=True)
    course_sheet = [0, wb['1 курс'], wb['2 курс'], wb['3 курс']]
    course_educational_programs_count = [0, 11, 9, 8]
    group_sheet = wb['Группы']


def set_schedule_workload() -> None:
    global schedule_wb
    schedule_wb = openpyxl.load_workbook(os.path.join(
        settings.MEDIA_ROOT, 'schedule.xlsx'), data_only=True)


def parse_subjects(course: int) -> list:
    subjects = []
    for i in range(10, course_sheet[course].max_row + 1):
        department = course_sheet[course].cell(
            row=i, column=course_educational_programs_count[course] + 14).value
        if isinstance(department, str) and functions.compare_strings(department, DepartmentEnum.computer_engineering.value):
            subject = Subject()
            subject.configure_subject(
                i, DepartmentEnum.computer_engineering.id, course_sheet[course], course_educational_programs_count[course])
            subjects.append(subject)
    return subjects


def parse_groups() -> list:
    groups = []
    for i in range(2, group_sheet.max_row + 1):
        group = Group()
        group.name = group_sheet.cell(row=i, column=1).value
        group.course = functions.clear_int(
            group_sheet.cell(row=i, column=2).value)
        group.number_of_students = functions.clear_int(
            group_sheet.cell(row=i, column=3).value)
        group.code = group.name.split('-')[0]
        groups.append(group)
    return groups


def parse_groups_for_subject(course: int, subject: Subject, allowed_group_codes) -> list:
    groups = []
    for i in range(2, len(allowed_group_codes) + 2):
        subject_name = course_sheet[course].cell(
            row=subject.excel_row_index, column=1).value
        if functions.compare_strings(subject.name, subject_name):
            trimester = functions.clear_int(course_sheet[course].cell(
                row=subject.excel_row_index, column=i).value)
            group_code = course_sheet[course].cell(row=8, column=i).value
            if group_code in allowed_group_codes and trimester > 0:
                groups.append((group_code, trimester))
    return groups


def parse_schedule() -> list:
    week_days = {'Monday', 'Tuesday', 'Wednesday',
                 'Thursday', 'Friday', 'Saturday'}
    schedule = []
    for group_sheet in schedule_wb:
        group_name = group_sheet.title
        group_schedule = collections.defaultdict(list)
        row = 2
        while row < group_sheet.max_row + 1:
            day_of_the_week = group_sheet.cell(row=row, column=1).value
            day_schedule = collections.defaultdict(list)
            if day_of_the_week not in week_days:
                break
            row += 1
            while group_sheet.cell(row=row, column=1).value is None:
                time = group_sheet.cell(row=row, column=2).value
                if time is None:
                    break
                subject_name = functions.formated(
                    group_sheet.cell(row=row, column=3).value)
                classroom = group_sheet.cell(row=row, column=4).value
                type = group_sheet.cell(row=row, column=5).value
                teacher = group_sheet.cell(row=row, column=6).value
                day_schedule[day_of_the_week].append(
                    [time, subject_name, classroom, type, teacher])
                row += 1
            group_schedule[group_name].append(day_schedule)
        schedule.append(group_schedule)
    return schedule
