from datetime import date
import os
import openpyxl
from openpyxl import Workbook
from api.service.save_service import create_excel_doc
from config.settings import ACADEMIC_LOAD_TEMPLATE_PATH, ACADEMIC_LOAD_PATH
from api.model.models import Workload
from accounts.models import Teacher
from openpyxl.styles import Border, Side, Font


wb = workload_sheet = None


def create_excel_workload() -> Workbook:
    global wb, workload_sheet
    create_excel_doc(ACADEMIC_LOAD_TEMPLATE_PATH, _excel_doc_path())
    wb = openpyxl.load_workbook(_excel_doc_path())
    workload_sheet = wb['Нагрузка']
    _populate_workload()
    wb.save(_excel_doc_path())
    return wb


def create_excel_workload_for_teacher(teacher_username: str) -> Workbook:
    global wb, workload_sheet
    teacher = Teacher.objects.get(username__exact=teacher_username)
    create_excel_doc(ACADEMIC_LOAD_TEMPLATE_PATH, _excel_doc_path_teacher(
        teacher.first_name, teacher.second_name))
    wb = openpyxl.load_workbook(_excel_doc_path_teacher(
        teacher.first_name, teacher.second_name))
    workload_sheet = wb['Нагрузка']
    _populate_workload_for_teacher(teacher, start_row_index=7)
    wb.save(_excel_doc_path_teacher(teacher.first_name, teacher.second_name))
    return wb


def _populate_workload() -> None:
    all_teacher = Teacher.objects.all()
    row_index = 7
    for teacher in all_teacher:
        row_index += _populate_workload_for_teacher(teacher, row_index)
    wb.save(_excel_doc_path())


def _populate_workload_for_teacher(teacher: Teacher, start_row_index: int) -> int:
    row_index = start_row_index
    teacher_workloads = Workload.objects.filter(
        teacher_id__exact=teacher.id).order_by('group_subject__subject__name', 'group_subject__group__name')
    if not teacher_workloads:
        return 0
    teacher_subjects = set(teacher_workloads.values_list(
        'group_subject__subject__name', 'group_subject__subject__office_hour'))
    start_row_index = row_index
    for subject_name, office_hour in teacher_subjects:
        workload_sheet.cell(row=row_index, column=16).value = office_hour
        teacher_lecture_workloads = teacher_workloads.filter(
            group_subject__subject__name__exact=subject_name, is_lecture__exact=True)
        teacher_practice_workloads = teacher_workloads.filter(
            group_subject__subject__name__exact=subject_name, is_practice__exact=True)
        teacher_lab_workloads = teacher_workloads.filter(
            group_subject__subject__name__exact=subject_name, is_lab__exact=True)
        if teacher_lecture_workloads:
            _configure_lecture_cells(teacher_lecture_workloads, row_index)
            row_index += 1
        if teacher_practice_workloads:
            _configure_cells(teacher_practice_workloads,
                             row_index, is_practice=True)
            row_index += len(teacher_practice_workloads)
        if teacher_lab_workloads:
            _configure_cells(teacher_lab_workloads, row_index, is_lab=True)
            row_index += len(teacher_lab_workloads)
    _merge_column(start_row_index, row_index)
    _configure_teacher_cell(teacher, start_row_index, row_index)
    _set_borders(row_index)
    row_index += 1
    return row_index - start_row_index


def _set_borders(row_index: int) -> None:
    other_side = Side(border_style='thin')
    bottom_side = Side(border_style='medium')
    for col in range(2, workload_sheet.max_column + 1):
        workload_sheet.cell(row=row_index, column=col).border = Border(
            bottom=bottom_side, top=other_side, left=other_side, right=other_side)


def _merge_column(start_row_index: int, end_row_index: int) -> None:
    for col in range(2, 7):
        workload_sheet.merge_cells(
            start_row=start_row_index, start_column=col, end_row=end_row_index, end_column=col)


def _configure_lecture_cells(workloads, row_index: int) -> None:
    _configure_workload_cell(workloads[0], row_index)
    lecture_groups = ', '.join(
        [workload.group_subject.group.name for workload in workloads])
    number_of_students = sum(
        [workload.group_subject.group.number_of_students for workload in workloads])
    workload_sheet.cell(row=row_index, column=9).value = lecture_groups
    workload_sheet.cell(row=row_index, column=11).value = number_of_students
    workload_sheet.cell(
        row=row_index, column=13).value = workloads[0].group_subject.subject.lecture_hour


def _configure_cells(workloads, row_index: int, is_practice: bool = False, is_lab: bool = False) -> None:
    for workload in workloads:
        _configure_workload_cell(workload, row_index, is_practice, is_lab)
        row_index += 1


def _get_sum_formula(start_row_index: int, end_row_index: int, column_index: int) -> str:
    return f"=SUM({workload_sheet.cell(row=start_row_index, column=column_index).coordinate}:{workload_sheet.cell(row=end_row_index, column=column_index).coordinate})"


def _set_sums(start_row_index: int, end_row_index: int) -> None:
    for i in range(13, 19):
        workload_sheet.cell(row=end_row_index, column=i).value = _get_sum_formula(
            start_row_index, end_row_index - 1, i)


def _set_fonts(row_index: int) -> None:
    for i in range(13, 21):
        workload_sheet.cell(row=row_index, column=i).font = Font(
            name="Times New Roman", size=10, bold=True)


def _configure_teacher_cell(teacher: Teacher, start_row_index: int, end_row_index: int) -> None:
    full_name = f"{teacher.first_name} {teacher.second_name}"
    workload_sheet.cell(row=start_row_index,
                        column=2).value = full_name
    workload_sheet.cell(row=start_row_index, column=3).value = teacher.position
    workload_sheet.cell(row=start_row_index, column=4).value = teacher.kpi
    workload_sheet.cell(row=start_row_index, column=5).value = teacher.one_rate
    workload_sheet.cell(row=start_row_index, column=6).value = teacher.load
    workload_sheet.cell(row=end_row_index,
                        column=19).value = teacher.total_hour
    workload_sheet.cell(
        row=end_row_index, column=20).value = round(teacher.total_hour / teacher.one_rate, 2)
    _set_sums(start_row_index, end_row_index)
    _set_fonts(end_row_index)


def _configure_workload_cell(workload: Workload, row_index: int, is_practice: bool = False, is_lab: bool = False) -> None:
    subject = workload.group_subject.subject
    group = workload.group_subject.group
    workload_sheet.cell(
        row=row_index, column=7).value = subject.name.capitalize()
    workload_sheet.cell(
        row=row_index, column=8).value = group.course
    workload_sheet.cell(row=row_index, column=9).value = group.name
    workload_sheet.cell(
        row=row_index, column=10).value = workload.group_subject.trimester
    workload_sheet.cell(
        row=row_index, column=11).value = group.number_of_students
    workload_sheet.cell(
        row=row_index, column=12).value = subject.credits
    if is_practice:
        workload_sheet.cell(
            row=row_index, column=15).value = subject.practice_hour
    if is_lab:
        workload_sheet.cell(
            row=row_index, column=14).value = subject.lab_hour


def _formatted_current_date() -> None:
    today = date.today()
    return date.strftime(today, "%d.%m.%Y")


def _excel_doc_path_teacher(first_name: str, second_name: str) -> None:
    directory = f"{ACADEMIC_LOAD_PATH}/{first_name} {second_name}"
    if not os.path.exists(directory):
        os.makedirs(directory)
    return f"{directory}/Нагрузка {_formatted_current_date()}.xlsx"


def _excel_doc_path() -> None:
    return f"{ACADEMIC_LOAD_PATH}/Нагрузка {_formatted_current_date()}.xlsx"
